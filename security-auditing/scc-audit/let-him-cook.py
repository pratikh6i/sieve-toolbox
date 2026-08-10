import os
import sys
import re
import glob
import traceback
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart3D, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.text import CharacterProperties, Font as DrawingFont


def parse_file_paths(input_str):
    """
    Parses space/comma-separated strings or wildcard patterns into a list of existing CSV file paths.
    """
    tokens = re.findall(r'"([^"]+)"|\'([^\']+)\'|(\S+)', input_str)
    raw_tokens = [t[0] or t[1] or t[2] for t in tokens]
    
    valid_files = []
    for token in raw_tokens:
        token_clean = token.strip(',;')
        if '*' in token_clean or '?' in token_clean:
            matched = glob.glob(token_clean)
            valid_files.extend([f for f in matched if os.path.isfile(f)])
        elif os.path.isfile(token_clean):
            valid_files.append(token_clean)

    # Remove duplicates while preserving user selection order
    seen = set()
    unique_files = []
    for f in valid_files:
        abs_f = os.path.abspath(f)
        if abs_f not in seen:
            seen.add(abs_f)
            unique_files.append(f)

    return unique_files


def prompt_user_inputs():
    """Interactively prompts for Customer Name and multiple CSV input file paths."""
    print("=" * 80)
    print("      GCP SECURITY COMMAND CENTER (SCC) EXECUTIVE REPORT GENERATOR      ")
    print("=" * 80)

    # Prompt Customer Name
    while True:
        customer_name = input("\nEnter Customer Name: ").strip()
        if customer_name:
            break
        print("[ERROR] Customer Name cannot be empty. Please try again.")

    # Prompt Multiple CSV File Paths
    selected_files = []
    print("\n------------------------------------------------------------------------")
    print("PROVIDE CSV FILE PATHS:")
    print("  • You can enter a single file path, multiple paths (space or comma separated),")
    print("    or wildcard patterns (e.g. *.csv or path/to/*.csv).")
    print("  • Enter additional paths or press Enter / type 'done' when finished.")
    print("------------------------------------------------------------------------")

    while True:
        if not selected_files:
            user_input = input("\nEnter CSV path(s) / pattern: ").strip()
        else:
            user_input = input("\nEnter more CSV path(s) (or press Enter if done): ").strip()

        if user_input.lower() in ['q', 'quit']:
            print("\n[INFO] Operation cancelled by user. Exiting.")
            sys.exit(0)

        if not user_input or user_input.lower() == 'done':
            if selected_files:
                break
            else:
                print("[ERROR] No valid CSV files provided yet. Please enter at least one file path.")
                continue

        parsed = parse_file_paths(user_input)
        if not parsed:
            print(f"[ERROR] No valid CSV files found matching: '{user_input}'. Please verify paths and try again.")
            continue

        for f in parsed:
            if f not in selected_files:
                selected_files.append(f)
                print(f"  [+] Added file: {f}")

    print("\n" + "-" * 60)
    print(f"SUMMARY OF {len(selected_files)} SELECTED CSV FILE(S) FOR COMBINED PROCESSING:")
    for idx, f in enumerate(selected_files, 1):
        print(f"  {idx}. {f}")
    print("-" * 60)

    return customer_name, selected_files


def load_and_classify_scc_data(file_paths):
    """Loads and concatenates multiple CSV files, classifying findings based on finding.category."""
    print(f"\n[REALTIME UPDATE] Reading and combining dataset from {len(file_paths)} CSV file(s)...")
    
    raw_dfs = []
    for fp in file_paths:
        try:
            df_temp = pd.read_csv(fp, low_memory=False)
            raw_dfs.append(df_temp)
            print(f"  --> Loaded '{os.path.basename(fp)}': {len(df_temp):,} records")
        except Exception as e:
            print(f"  [WARNING] Could not read file '{fp}': {e}")

    if not raw_dfs:
        raise ValueError("Could not extract data from any of the provided CSV file paths.")

    df_combined = pd.concat(raw_dfs, ignore_index=True)
    total_records = len(df_combined)
    print(f"[REALTIME UPDATE] Total combined findings across all CSV files: {total_records:,}")

    # Standardize category column lookup
    category_col = None
    for col in df_combined.columns:
        if col.lower() == 'finding.category':
            category_col = col
            break

    if not category_col:
        raise ValueError("Uploaded CSV files are missing required column: 'finding.category'")

    # Report category breakdown across combined dataset
    category_counts = df_combined[category_col].astype(str).str.upper().value_counts()
    print("\n" + "-" * 60)
    print("COMBINED FINDINGS CATEGORY BREAKDOWN:")
    print("-" * 60)
    for cat, count in category_counts.items():
        print(f"  • {cat}: {count:,} records")
    print("-" * 60)

    # Primary Category Filtering: OS_VULNERABILITY vs SOFTWARE_VULNERABILITY
    os_mask = df_combined[category_col].astype(str).str.upper() == 'OS_VULNERABILITY'
    sw_mask = df_combined[category_col].astype(str).str.upper() == 'SOFTWARE_VULNERABILITY'

    other_mask = ~(os_mask | sw_mask)
    other_df = df_combined[other_mask]

    if not other_df.empty:
        other_categories = other_df[category_col].astype(str).str.upper().value_counts()
        print("\n[NOTICE] Findings other than OS and Software detected in combined input:")
        for cat, count in other_categories.items():
            print(f"  --> Excluded Category '{cat}': {count:,} findings")
        print("[NOTICE] These non-OS/Software findings are IGNORED in the final Excel reports.")

    # Process records
    def extract_structured_records(sub_df):
        records = []
        for idx, row in sub_df.iterrows():
            # Project Name
            proj = row.get('resource.gcp_metadata.project_display_name')
            if pd.isna(proj) or not str(proj).strip():
                proj_str = str(row.get('resource.gcp_metadata.project', ''))
                proj_match = re.search(r'projects/([^/]+)', proj_str)
                proj = proj_match.group(1) if proj_match else "Unknown_Project"

            # Resource Display Name
            res_name = row.get('resource.display_name')
            if pd.isna(res_name) or not str(res_name).strip():
                res_str = str(row.get('resource.name', ''))
                res_match = re.search(r'instances/([^/,\s]+)', res_str)
                res_name = res_match.group(1) if res_match else "Unknown_Resource"

            # Severity
            sev = str(row.get('finding.severity', 'HIGH')).upper()
            if sev not in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
                sev = 'HIGH'

            # State & Category
            state = str(row.get('finding.state', 'ACTIVE')).upper()
            cat = str(row.get(category_col, 'VULNERABILITY')).upper()

            # CVE ID
            cve = str(row.get('finding.vulnerability.cve.id', 'N/A'))
            if cve.lower() == 'nan' or not cve.strip():
                cve = 'N/A'

            # Resource Type
            res_type = str(row.get('resource.type', 'google.compute.Instance'))
            if res_type.lower() == 'nan':
                res_type = 'google.compute.Instance'

            # Component / Package Name
            offending_pkg = str(row.get('finding.vulnerability.offending_package.package_name', ''))
            fixed_pkg = str(row.get('finding.vulnerability.fixed_package.package_name', ''))
            
            pkg_name = fixed_pkg if fixed_pkg and fixed_pkg.lower() != 'nan' else offending_pkg
            if not pkg_name or pkg_name.lower() == 'nan':
                desc = str(row.get('finding.description', '')).lower()
                if 'kernel' in desc: pkg_name = 'kernel'
                elif 'openssh' in desc: pkg_name = 'openssh'
                elif 'bash' in desc: pkg_name = 'bash'
                else: pkg_name = 'system-package'

            records.append({
                'finding_id': row.get('finding.name', f'finding_{idx}'),
                'project': proj,
                'resource_name': res_name,
                'severity': sev,
                'state': state,
                'category': cat,
                'cve_id': cve,
                'resource_type': res_type,
                'package_type': pkg_name
            })
        return pd.DataFrame(records)

    df_os = extract_structured_records(df_combined[os_mask])
    df_sw = extract_structured_records(df_combined[sw_mask])

    return df_os, df_sw


def build_excel_workbook(df, report_title_prefix, customer_name, output_filepath):
    """Builds formatted 5-tab workbook with embedded Google Sheets compatible charts."""
    wb = openpyxl.Workbook()

    # Reusable Formatting Objects
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Blue [#0000FF] Header Fill
    header_font = Font(name="Google Sans", size=11, bold=True, color="FFFFFF") # Plain White Text
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    def format_title_block(ws, title_text):
        ws.views.sheetView[0].showGridLines = True
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = title_text
        title_cell.font = Font(name="Google Sans", size=15, bold=True, color="000000")
        
        ws["A2"] = f"Customer: {customer_name} | Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Source: GCP Security Command Center"
        ws["A2"].font = Font(name="Google Sans", size=9, italic=True, color="595959")

    def apply_chart_font(chart, font_name="Google Sans"):
        cp = CharacterProperties(latin=DrawingFont(typeface=font_name))
        if hasattr(chart, 'title') and chart.title and hasattr(chart.title, 'tx') and chart.title.tx and chart.title.tx.rich:
            for p in chart.title.tx.rich.p:
                p.pPr.defRPr = cp

    def style_table(ws, start_row, headers, df_data):
        # Write Headers
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write Rows
        current_row = start_row + 1
        for r_idx, row_data in enumerate(df_data):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                cell.font = Font(name="Google Sans", size=10)
                cell.border = thin_border
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
            current_row += 1

        # Column Width Adjustment
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 2 and cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

        return current_row

    # -------------------------------------------------------------------------
    # TAB 1: EXECUTIVE SUMMARY
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    format_title_block(ws1, f"Google Cloud SCC — {report_title_prefix} Summary")

    total = len(df)
    active = len(df[df['state'] == 'ACTIVE'])
    critical = len(df[df['severity'] == 'CRITICAL'])
    high = len(df[df['severity'] == 'HIGH'])

    ws1["A4"] = "TOTAL FINDINGS"; ws1["B4"] = total
    ws1["A5"] = "ACTIVE FINDINGS"; ws1["B5"] = active
    ws1["A6"] = "CRITICAL SEVERITY"; ws1["B6"] = critical
    ws1["A7"] = "HIGH SEVERITY"; ws1["B7"] = high

    for r in range(4, 8):
        ws1[f"A{r}"].font = Font(name="Google Sans", size=10, bold=True)
        ws1[f"B{r}"].font = Font(name="Google Sans", size=11, bold=True)
        ws1[f"B{r}"].alignment = Alignment(horizontal="center")

    sev_df = df.groupby('severity').agg(
        Total_Count=('finding_id', 'count'),
        Active=('state', lambda x: (x == 'ACTIVE').sum())
    ).reset_index()

    style_table(ws1, 10, ["Severity Level", "Total Count", "Active Findings"], sev_df.values.tolist())

    # 3D Pie Chart for Severity
    pie = PieChart3D()
    labels = Reference(ws1, min_col=1, min_row=11, max_row=10 + len(sev_df))
    chart_data = Reference(ws1, min_col=2, min_row=10, max_row=10 + len(sev_df))
    pie.add_data(chart_data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Total Count"
    pie.width = 15; pie.height = 7.5

    # REQUIREMENT 3: Visible slice count/value on pie chart
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True

    # REQUIREMENT: Color Slices (Red [#CC0000] for Critical, Yellow [#F1C232] for High)
    if len(pie.series) > 0:
        for idx, row in sev_df.iterrows():
            sev = str(row['severity']).upper()
            dp = DataPoint(idx=idx)
            if sev == 'CRITICAL':
                dp.graphicalProperties.solidFill = "CC0000" # Red
            elif sev == 'HIGH':
                dp.graphicalProperties.solidFill = "F1C232" # Yellow
            elif sev == 'MEDIUM':
                dp.graphicalProperties.solidFill = "5BC0DE"
            else:
                dp.graphicalProperties.solidFill = "5CB85C"
            pie.series[0].data_points.append(dp)

    # REQUIREMENT 2: Google Sans Font
    apply_chart_font(pie, "Google Sans")
    ws1.add_chart(pie, "D2")

    # -------------------------------------------------------------------------
    # TAB 2: PROJECT DISTRIBUTION (STACKED COLUMN CHART)
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Project Distribution")
    format_title_block(ws2, f"{report_title_prefix} — Project Distribution")

    proj_summary = df.groupby('project').agg(
        Total_Findings=('finding_id', 'count'),
        Critical_Count=('severity', lambda x: (x == 'CRITICAL').sum()),
        High_Count=('severity', lambda x: (x == 'HIGH').sum())
    ).reset_index().sort_values(by=['Critical_Count', 'Total_Findings'], ascending=False)

    style_table(ws2, 4, ["GCP Project Name", "Total Findings", "Critical Severity", "High Severity"], proj_summary.values.tolist())

    # REQUIREMENT 1: Stacked Column Chart for Sheet 2
    bar_proj = BarChart()
    bar_proj.type = "col"        # Vertical Columns
    bar_proj.grouping = "stacked" # Stacked Grouping
    bar_proj.overlap = 100
    bar_proj.title = "Total Findings, Critical Severity and High Severity"
    bar_proj.y_axis.title = "Count"
    bar_proj.x_axis.title = "GCP Project Name"

    # Data series range: Critical Severity (Col C) and High Severity (Col D)
    bar_data = Reference(ws2, min_col=3, min_row=4, max_col=4, max_row=4 + len(proj_summary))
    bar_cats = Reference(ws2, min_col=1, min_row=5, max_row=4 + len(proj_summary))
    bar_proj.add_data(bar_data, titles_from_data=True)
    bar_proj.set_categories(bar_cats)
    
    # Custom Series Fill Colors: Critical = Red [#CC0000], High = Yellow [#F1C232]
    if len(bar_proj.series) >= 1:
        bar_proj.series[0].graphicalProperties.solidFill = "CC0000" # Red for Critical
    if len(bar_proj.series) >= 2:
        bar_proj.series[1].graphicalProperties.solidFill = "F1C232" # Yellow for High

    bar_proj.width = 16; bar_proj.height = 9.5
    # REQUIREMENT 2: Google Sans Font
    apply_chart_font(bar_proj, "Google Sans")
    ws2.add_chart(bar_proj, "F2")

    # -------------------------------------------------------------------------
    # TAB 3: RESOURCE ANALYSIS
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet(title="Resource Analysis")
    format_title_block(ws3, f"{report_title_prefix} — Resource Analysis")

    res_summary = df.groupby(['resource_name', 'project']).agg(
        Total_Findings=('finding_id', 'count'),
        Critical=('severity', lambda x: (x == 'CRITICAL').sum()),
        High=('severity', lambda x: (x == 'HIGH').sum())
    ).reset_index().sort_values(by='Total_Findings', ascending=False)

    last_r3 = style_table(ws3, 4, ["Resource Display Name", "GCP Project", "Total Findings", "Critical", "High"], res_summary.values.tolist())

    type_summary = df.groupby('resource_type').size().reset_index(name='Count')
    ws3.cell(row=last_r3 + 2, column=1, value="Resource Type Breakdown").font = Font(bold=True, size=11, color="000000")
    style_table(ws3, last_r3 + 3, ["Resource Type", "Findings Count"], type_summary.values.tolist())

    # -------------------------------------------------------------------------
    # TAB 4: PROJECT X RESOURCE MATRIX
    # -------------------------------------------------------------------------
    ws4 = wb.create_sheet(title="Project x Resource Matrix")
    format_title_block(ws4, f"{report_title_prefix} — Project & Resource Matrix")

    combined_summary = df.groupby(['project', 'resource_name', 'resource_type']).agg(
        Total_Findings=('finding_id', 'count'),
        Critical=('severity', lambda x: (x == 'CRITICAL').sum()),
        High=('severity', lambda x: (x == 'HIGH').sum())
    ).reset_index().sort_values(by=['project', 'Total_Findings'], ascending=False)

    style_table(ws4, 4, ["GCP Project", "Resource Display Name", "Resource Type", "Findings Count", "Critical", "High"], combined_summary.values.tolist())

    # -------------------------------------------------------------------------
    # TAB 5: COMPONENTS & CVES
    # -------------------------------------------------------------------------
    ws5 = wb.create_sheet(title="Components & CVEs")
    format_title_block(ws5, f"{report_title_prefix} — Components & CVE Breakdown")

    pkg_summary = df.groupby('package_type').size().reset_index(name='Findings_Count').sort_values(by='Findings_Count', ascending=False)
    last_r5 = style_table(ws5, 4, ["Impacted Component / Package", "Findings Count"], pkg_summary.values.tolist())

    cve_summary = df[df['cve_id'] != 'N/A'].groupby('cve_id').agg(
        Count=('finding_id', 'count'),
        Severity=('severity', 'first'),
        Sample_Resource=('resource_name', 'first')
    ).reset_index().sort_values(by='Count', ascending=False)

    ws5.cell(row=last_r5 + 2, column=1, value="Top CVE Summary").font = Font(bold=True, size=11, color="000000")
    style_table(ws5, last_r5 + 3, ["CVE ID", "Finding Count", "Severity", "Sample Impacted Resource"], cve_summary.values.tolist())

    # Top Packages Chart
    pie_pkg = PieChart3D()
    pkg_labels = Reference(ws5, min_col=1, min_row=5, max_row=min(18, 4 + len(pkg_summary)))
    pkg_data = Reference(ws5, min_col=2, min_row=4, max_row=min(18, 4 + len(pkg_summary)))
    pie_pkg.add_data(pkg_data, titles_from_data=True)
    pie_pkg.set_categories(pkg_labels)
    pie_pkg.title = "Findings Count vs. Impacted Component / Package"
    pie_pkg.width = 15; pie_pkg.height = 7.5

    # REQUIREMENT 3: Visible slice count/value on pie chart
    pie_pkg.dataLabels = DataLabelList()
    pie_pkg.dataLabels.showVal = True

    # REQUIREMENT 2: Google Sans Font
    apply_chart_font(pie_pkg, "Google Sans")
    ws5.add_chart(pie_pkg, "D1")

    # Save Excel Workbook
    wb.save(output_filepath)
    print(f"[REALTIME UPDATE] Successfully generated report: '{output_filepath}'")


def main():
    try:
        customer_name, file_paths = prompt_user_inputs()
        
        # Sanitize customer name for output file naming
        clean_cust_name = re.sub(r'[^\w\s-]', '', customer_name).strip().replace(' ', '_')
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        df_os, df_sw = load_and_classify_scc_data(file_paths)

        print("\n" + "=" * 80)
        print("          GENERATING EXECUTIVE INSIGHT EXCEL WORKBOOKS          ")
        print("=" * 80)

        # 1) OS Vulnerability Findings Workbook
        if not df_os.empty:
            os_filename = f"{clean_cust_name}_OS_Vulnerability_Report_{timestamp}.xlsx"
            print(f"\n[REALTIME UPDATE] Building OS Findings Workbook ({len(df_os):,} findings)...")
            build_excel_workbook(df_os, "OS Vulnerability Findings", customer_name, os_filename)
        else:
            print("\n[INFO] No OS Vulnerability findings detected across provided files. Skipping OS report.")

        # 2) Software Vulnerability Findings Workbook
        if not df_sw.empty:
            sw_filename = f"{clean_cust_name}_Software_Vulnerability_Report_{timestamp}.xlsx"
            print(f"\n[REALTIME UPDATE] Building Software Findings Workbook ({len(df_sw):,} findings)...")
            build_excel_workbook(df_sw, "Software & App Vulnerability Findings", customer_name, sw_filename)
        else:
            print("\n[INFO] No Software Vulnerability findings detected across provided files. Skipping Software report.")

        print("\n" + "=" * 80)
        print("[SUCCESS] All tasks completed! Final Excel reports generated successfully.")
        print("=" * 80 + "\n")

    except Exception as e:
        print("\n" + "!" * 80)
        print("[FATAL ERROR OCCURRED] An unexpected error stopped program execution.")
        print(f"  Error Type: {type(e).__name__}")
        print(f"  Error Message: {e}")
        print("\n--- DETAILED TECHNICAL TRACEBACK (FOR DEVELOPER / AI DEBUGGING) ---")
        traceback.print_exc()
        print("!" * 80 + "\n")
        sys.exit(1)


if __name__ == "__main__":
    main() 
